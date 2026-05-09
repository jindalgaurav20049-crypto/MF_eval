from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from fpdf import FPDF
from openpyxl import Workbook


def build_fund_summary_excel(summary: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fund Summary"
    sheet.append(["Field", "Value"])
    for key, value in summary.items():
        sheet.append([key, _format_value(value)])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_fund_summary_pdf(summary: dict[str, Any]) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    pdf.cell(0, 10, txt="Fund Summary", ln=1)
    for key, value in summary.items():
        pdf.cell(0, 8, txt=f"{key}: {_format_value(value)}", ln=1)
    return pdf.output(dest="S").encode("latin-1")


def build_portfolio_excel(entries: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Portfolio"
    if not entries:
        sheet.append(["No portfolio data"])
    else:
        headers = list(entries[0].keys())
        sheet.append(headers)
        for entry in entries:
            sheet.append([_format_value(entry.get(h)) for h in headers])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_portfolio_pdf(entries: list[dict[str, Any]]) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    pdf.cell(0, 10, txt="Portfolio Export", ln=1)
    for entry in entries:
        line = ", ".join(f"{k}: {_format_value(v)}" for k, v in entry.items())
        pdf.multi_cell(0, 8, txt=line)
    return pdf.output(dest="S").encode("latin-1")


def _format_value(value: Any) -> str:
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return "-"
    return str(value)
